import sys
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QVBoxLayout, QHBoxLayout, QLabel, QSizePolicy, QMessageBox
from PyQt5.QtCore import Qt, QDateTime
from PyQt5.QtGui import QFont, QFontDatabase

import openpyxl

class RadioButtonExample(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('고성시장 토요장터 방명록: 제작_서성현')
        self.resize(800, 600)

        layout = QVBoxLayout()
        self.addTitleLabel(layout)
        layout.addSpacing(30)
        self.addGroup1Buttons(layout)
        layout.addSpacing(30)
        self.addGroup2Buttons(layout)
        layout.addSpacing(50)
        self.addSubmitButton(layout)

        self.setLayout(layout)

        self.selected_buttons = {group1_button: None for group1_button in self.group1_buttons}
        self.selected_buttons.update({group2_button: None for group2_button in self.group2_buttons})

        self.excel_file_selected = False

    def addTitleLabel(self, layout):
        title_label = QLabel('간단 방문등록')
        title_label.setAlignment(Qt.AlignCenter)
        title_font = QFont('Arial', 100)
        title_label.setFont(title_font)
        layout.addWidget(title_label)

    def addGroup1Buttons(self, layout):
        group1_label = QLabel('연령:')
        group_font = QFont('Arial', 40)
        group1_label.setFont(group_font)
        layout.addWidget(group1_label)

        group1_layout = QHBoxLayout()
        self.group1_buttons = []
        group1_button_names = ['10대미만', '10,20대', '30대', '40대', '50대이상']
        for name in group1_button_names:
            button = QPushButton(name)
            button.setCheckable(True)
            button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            button.clicked.connect(self.onButtonClicked)
            group1_layout.addWidget(button)
            self.group1_buttons.append(button)

        layout.addLayout(group1_layout)

    def addGroup2Buttons(self, layout):
        group2_label = QLabel('지역:')
        group_font = QFont('Arial', 40)
        group2_label.setFont(group_font)
        layout.addWidget(group2_label)

        group2_layout = QHBoxLayout()
        self.group2_buttons = []
        group2_button_names = ['고성', '경남', '그외']
        for name in group2_button_names:
            button = QPushButton(name)
            button.setCheckable(True)
            button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            button.clicked.connect(self.onButtonClicked)
            group2_layout.addWidget(button)
            self.group2_buttons.append(button)

        layout.addLayout(group2_layout)

    def addSubmitButton(self, layout):
        submit_button = QPushButton('완료')
        submit_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        submit_button.clicked.connect(self.onSubmitButtonClicked)

        font_id = QFontDatabase.addApplicationFont("NanumBarunGothic.ttf")
        family = QFontDatabase.applicationFontFamilies(font_id)[0]
        font = QFont(family, 50)
        submit_button.setFont(font)

        layout.addWidget(submit_button)

    def onButtonClicked(self):
        sender = self.sender()
        if sender in self.group1_buttons:
            for button in self.group1_buttons:
                button.setChecked(button is sender)
            self.selected_buttons[sender] = sender.text()

        elif sender in self.group2_buttons:
            for button in self.group2_buttons:
                button.setChecked(button is sender)
            self.selected_buttons[sender] = sender.text()

    def onSubmitButtonClicked(self):
        current_time = QDateTime.currentDateTime()
        print("제출 버튼이 클릭되었습니다.")
        print("방문자 번호:", self.visitor_number)

        if not self.isOneButtonSelected(self.group1_buttons) or not self.isOneButtonSelected(self.group2_buttons):
            QMessageBox.warning(self, "경고", "각 그룹에서 하나씩 선택해주세요.", QMessageBox.Ok)
            return

        self.excel_file_selected = True

        visitor_data = [
            self.visitor_number,
            self.getCheckedButtonText(self.group1_buttons),
            self.getCheckedButtonText(self.group2_buttons),
            current_time.toString(Qt.DefaultLocaleLongDate),
        ]

        try:
            wb = openpyxl.load_workbook(self.excel_file_name)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            wb.create_sheet(index=0)

        sheet = wb.active
        sheet.append(visitor_data)

        wb.save(self.excel_file_name)
        print("선택된 방문자 데이터가 엑셀 파일에 기록되었습니다.")

        self.visitor_number += 1

        self.resetSelectedButtons()

    def getCheckedButtonText(self, buttons):
        for button in buttons:
            if button.isChecked():
                return button.text()
        return None

    def isOneButtonSelected(self, buttons):
        return sum(button.isChecked() for button in buttons) == 1

    def resetSelectedButtons(self):
        for button in self.group1_buttons + self.group2_buttons:
            button.setChecked(False)
            self.selected_buttons[button] = None

if __name__ == '__main__':
    app = QApplication(sys.argv)
