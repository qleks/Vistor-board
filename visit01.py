import sys
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QVBoxLayout, QHBoxLayout, QLabel, QSizePolicy, QMessageBox
from PyQt5.QtCore import Qt, QDateTime, QSize
from PyQt5.QtGui import QFont, QColor, QResizeEvent
from PyQt5.QtMultimedia import QSound
import qt_material
import openpyxl

class RadioButtonExample(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.setWindowTitle('고성시장 토요장터 방명록: 제작_서성현')
        self.resize(800, 600)

        layout = QVBoxLayout()

        # 그룹 간 간격 추가
        layout.addSpacing(30)

        # 제목 추가
        self.title_label = QLabel('토요장터 방문등록')
        self.title_label.setAlignment(Qt.AlignCenter)
        self.title_font = QFont()
        self.title_font.setPointSize(self.calculateFontSize())
        self.title_label.setFont(self.title_font)
        layout.addWidget(self.title_label)

        # 그룹 간 간격 추가
        layout.addSpacing(30)

        # 방문자 번호를 저장하기 위한 변수 초기화
        self.visitor_number = self.getNextVisitorNumber()

        # 그룹 1의 버튼 생성
        group1_label = QLabel('연령:')
        group_font = QFont()
        group_font.setPointSize(50)
        group1_label.setFont(group_font)
        layout.addWidget(group1_label)

        group1_layout = QHBoxLayout()
        self.group1_buttons = []
        group1_button_names = ['10대미만', '10,20대', '30대', '40대', '50대이상']
        for name in group1_button_names:
            button = QPushButton(name)
            button.setCheckable(True)
            button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            button.setFont(QFont("Arial", 30))
            button.setStyleSheet("QPushButton { border: 2px solid black; color: black; }")
            button.clicked.connect(self.onButtonClicked)
            group1_layout.addWidget(button)
            self.group1_buttons.append(button)

        layout.addLayout(group1_layout)

        # 그룹 간 간격 추가
        layout.addSpacing(30)

        # 그룹 2의 버튼 생성
        group2_label = QLabel('지역:')
        group2_label.setFont(group_font)
        layout.addWidget(group2_label)

        group2_layout = QHBoxLayout()
        self.group2_buttons = []
        group2_button_names = ['고성', '경남', '그외']
        for name in group2_button_names:
            button = QPushButton(name)
            button.setCheckable(True)
            button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            button.setFont(QFont("Arial", 30))
            button.setStyleSheet("QPushButton { border: 2px solid black; color: black; }")
            button.clicked.connect(self.onButtonClicked)
            group2_layout.addWidget(button)
            self.group2_buttons.append(button)

        layout.addLayout(group2_layout)

        # 그룹 간 간격 추가
        layout.addSpacing(50)

        # "제출" 버튼 생성
        submit_button = QPushButton('제출하기')
        submit_button.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        submit_button.setFont(QFont("Arial", 30))
        submit_button.setStyleSheet("QPushButton { border: 2px solid black; color: black; }")
        submit_button.clicked.connect(self.onSubmitButtonClicked)
        layout.addWidget(submit_button)

        self.setLayout(layout)

        # 선택된 버튼을 저장하는 변수 초기화
        self.selected_buttons = {group1_button: None for group1_button in self.group1_buttons}
        self.selected_buttons.update({group2_button: None for group2_button in self.group2_buttons})

        # 엑셀 파일 선택 여부를 나타내는 변수
        self.excel_file_selected = False

    def getNextVisitorNumber(self):
        self.excel_file_name = "고성시장_토요장터.xlsx"
        try:
            wb = openpyxl.load_workbook(self.excel_file_name)
            sheet = wb.active
            last_row = sheet.max_row
            if last_row > 1:
                return sheet.cell(row=last_row, column=1).value + 1
            else:
                return 1
        except FileNotFoundError:
            return 1

    def onButtonClicked(self):
        sender = self.sender()
        if sender in self.group1_buttons:
            # 그룹 1 버튼 중 하나만 선택되도록 처리
            for button in self.group1_buttons:
                button.setChecked(button is sender)
            self.selected_buttons[sender] = sender.text()

        elif sender in self.group2_buttons:
            # 그룹 2 버튼 중 하나만 선택되도록 처리
            for button in self.group2_buttons:
                button.setChecked(button is sender)
            self.selected_buttons[sender] = sender.text()

    def onSubmitButtonClicked(self):
        current_time = QDateTime.currentDateTime()
        print("제출 버튼이 클릭되었습니다.")
        print("방문자 번호:", self.visitor_number)

        # Check if one button is selected in each group
        if not self.isOneButtonSelected(self.group1_buttons) or not self.isOneButtonSelected(self.group2_buttons):
            QMessageBox.warning(self, "경고", "각 그룹에서 하나씩 선택해주세요.", QMessageBox.Ok)
            return

        self.excel_file_selected = True  # Indicate that the file is already specified

        # 선택된 방문자 데이터를 엑셀 파일에 추가
        visitor_data = [
            self.visitor_number,
            self.getCheckedButtonText(self.group1_buttons),
            self.getCheckedButtonText(self.group2_buttons),
            current_time.toString(Qt.DefaultLocaleLongDate),
        ]

        try:
            wb = openpyxl.load_workbook(self.excel_file_name)
        except FileNotFoundError:
            wb = openpyxl.Workbook()
            wb.remove(wb.active)
            wb.create_sheet(index=0)

        sheet = wb.active
        sheet.append(visitor_data)

        wb.save(self.excel_file_name)
        print("선택된 방문자 데이터가 엑셀 파일에 기록되었습니다.")

        # Play notification sound
        QSound.play("notification.wav")

        # 방문자 번호 증가
        self.visitor_number += 1

        # Reset selected buttons after submitting
        self.resetSelectedButtons()

    def getCheckedButtonText(self, buttons):
        for button in buttons:
            if button.isChecked():
                return button.text()
        return None

    def isOneButtonSelected(self, buttons):
        return sum(button.isChecked() for button in buttons) == 1

    def resetSelectedButtons(self):
        # 모든 그룹의 버튼을 초기화
        for button in self.group1_buttons + self.group2_buttons:
            button.setChecked(False)
            self.selected_buttons[button] = None

    def resizeEvent(self, event: QResizeEvent):
        super().resizeEvent(event)
        self.title_font.setPointSize(self.calculateFontSize())
        self.title_label.setFont(self.title_font)

    def calculateFontSize(self):
        width = self.width()
        height = self.height()
        return min(width, height) // 10

if __name__ == '__main__':
    app = QApplication(sys.argv)
    qt_material.apply_stylesheet(app, theme='light_teal.xml')  # 밝은색 테마 적용
    window = RadioButtonExample()
    window.show()
    sys.exit(app.exec_())
